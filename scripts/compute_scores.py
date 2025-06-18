import os
import warnings
warnings.simplefilter('ignore', FutureWarning)

import yfinance as yf
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

INDIAN_TICKERS = [
    "ATUL.NS", "BAYERCROP.NS", "BATAINDIA.NS", "GRAPHITE.NS",
    "CARBORUNIV.NS", "FINCABLES.NS", "CEATLTD.NS", "ELGIEQUIP.NS",
    "HAPPSTMNDS.NS", "SKFINDIA.NS", "CYIENT.NS", "EIHOTEL.NS",
    "BLUEDART.NS", "ASTRAZEN.NS", "CESC.NS", "DEEPAKFERT.NS",
    "CCL.NS", "EIDPARRY.NS", "MGL.NS", "CERA.NS"
]

Z_OUTLIERS = []
RATIO_WARNINGS = []

def extract_scalar(val):
    if isinstance(val, pd.Series):
        return val.iloc[-1] if not val.empty else None
    return None if isinstance(val, pd.DataFrame) else val

def compute_z_score(bs, isf, ticker=None, year=None):
    try:
        wc = extract_scalar(bs.loc["Working Capital"]) if "Working Capital" in bs.index else None
        ta = extract_scalar(bs.loc["Total Assets"])
        re = extract_scalar(bs.loc["Retained Earnings"])
        tl = extract_scalar(bs.loc["Total Liabilities Net Minority Interest"])
        ebit = extract_scalar(isf.loc["EBIT"])
        if None in (wc, ta, re, tl, ebit) or ta == 0 or tl == 0:
            return None

        bve = ta - tl
        x1 = wc / ta
        x2 = re / ta
        x3 = ebit / ta
        x4 = bve / tl

        if abs(x1) > 1 or abs(x2) > 2 or abs(x3) > 1 or x4 > 10:
            RATIO_WARNINGS.append({
                "Ticker": ticker, "Year": year,
                "x1": round(x1, 4), "x2": round(x2, 4),
                "x3": round(x3, 4), "x4": round(x4, 4)
            })

        z = 3.25 + 6.56 * x1 + 3.26 * x2 + 6.72 * x3 + 1.05 * x4

        if z > 25:
            Z_OUTLIERS.append({
                "Ticker": ticker, "Year": year,
                "Z": round(z, 2), "x4": round(x4, 4)
            })

        return round(z, 2)
    except Exception:
        return None

def compute_piotroski_f_score(tk):
    fin, bs, cf = tk.financials, tk.balance_sheet, tk.cashflow
    if fin.shape[1] < 2 or bs.shape[1] < 2 or cf.shape[1] < 2:
        return None
    cols = fin.columns[:2]
    def val(df, row, col): return extract_scalar(df.loc[row, col]) if row in df.index else None

    try:
        roi0 = val(fin, "Net Income", cols[0]) / val(bs, "Total Assets", cols[0])
        roi1 = val(fin, "Net Income", cols[1]) / val(bs, "Total Assets", cols[1])
        cfo0 = val(cf, "Total Cash From Operating Activities", cols[0])
        ni0 = val(fin, "Net Income", cols[0])
        dl0, dl1 = val(bs, "Long Term Debt", cols[0]) or 0, val(bs, "Long Term Debt", cols[1]) or 0
        ca0, cl0 = val(bs, "Current Assets", cols[0]) or 0, val(bs, "Current Liabilities", cols[0]) or 0
        ca1, cl1 = val(bs, "Current Assets", cols[1]) or 0, val(bs, "Current Liabilities", cols[1]) or 0
        ce0, ce1 = val(bs, "Common Stock Equity", cols[0]), val(bs, "Common Stock Equity", cols[1])
        gm0 = val(fin, "Gross Profit", cols[0]) / val(fin, "Total Revenue", cols[0])
        gm1 = val(fin, "Gross Profit", cols[1]) / val(fin, "Total Revenue", cols[1])
        at0 = val(fin, "Total Revenue", cols[0]) / val(bs, "Total Assets", cols[0])
        at1 = val(fin, "Total Revenue", cols[1]) / val(bs, "Total Assets", cols[1])

        score = sum([
            roi0 > 0 if roi0 else 0,
            cfo0 > 0 if cfo0 else 0,
            roi0 > roi1 if roi0 and roi1 else 0,
            cfo0 > ni0 if cfo0 and ni0 else 0,
            (dl0 / cl0) < (dl1 / cl1) if cl0 and cl1 else 0,
            (ca0 / cl0) > (ca1 / cl1) if ca0 and cl0 and ca1 and cl1 else 0,
            ce0 <= ce1 if ce0 and ce1 else 0,
            gm0 > gm1 if gm0 and gm1 else 0,
            at0 > at1 if at0 and at1 else 0
        ])
        return score
    except:
        return None

def process_ticker(ticker):
    tk = yf.Ticker(ticker)
    try:
        sector = tk.info.get("sector", "Unknown")
    except:
        sector = "Unknown"

    fin, bs, cf = tk.financials, tk.balance_sheet, tk.cashflow
    if fin.empty or bs.empty:
        return pd.DataFrame()

    records = []
    for year in bs.columns[:5]:
        yr = year.year
        rec = {"Ticker": ticker, "Year": yr, "Sector": sector}
        if sector in ["Technology", "IT Services", "Software"]:
            score = compute_piotroski_f_score(tk)
            status = ("Strong" if score >= 6 else "Neutral" if score >= 4 else "Weak") if score is not None else None
            rec.update(Model="Piotroski F‑Score", Score=score, Status=status)
        elif sector in ["Finance", "Banks", "Insurance"]:
            rec.update(Model="Skipped", Score=None, Status="Finance Sector")
        else:
            z = compute_z_score(bs[year], fin[year], ticker=ticker, year=yr)
            status = "Distress" if z is not None and z < 1.1 else "Grey" if z is not None and z <= 2.6 else "Safe"
            rec.update(Model="Altman Z″", Score=z, Status=status)
        records.append(rec)

    return pd.DataFrame(records)

def highlight_excel(filepath):
    wb = load_workbook(filepath)
    ws = wb.active
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    for row in ws.iter_rows(min_row=2):
        if str(row[4].value).lower() in ["distress", "weak"]:
            for cell in row:
                cell.fill = red
    wb.save(filepath)

def plot_ztrends(df):
    os.makedirs("charts", exist_ok=True)
    for ticker, grp in df[df["Model"] == "Altman Z″"].groupby("Ticker"):
        plt.figure()
        plt.plot(grp["Year"], grp["Score"], marker='o')
        plt.axhline(1.1, color='red', linestyle='--', label='Distress')
        plt.axhline(2.6, color='orange', linestyle='--', label='Grey Zone')
        plt.title(f"{ticker} Z″ Trend")
        plt.xlabel("Year"); plt.ylabel("Z″ Score")
        plt.legend(); plt.grid(True)
        plt.savefig(f"charts/{ticker}_z_trend.png")
        plt.close()

def main():
    results = [process_ticker(t) for t in INDIAN_TICKERS]
    df = pd.concat([d for d in results if not d.empty], ignore_index=True)
    df.sort_values(["Ticker", "Year"], inplace=True)

    df.to_csv("indian_smart_z_fscore.csv", index=False)
    df.to_excel("indian_smart_z_fscore.xlsx", index=False)
    highlight_excel("indian_smart_z_fscore.xlsx")
    plot_ztrends(df)

    # ✅ Always export diagnostics now
    pd.DataFrame(Z_OUTLIERS).to_csv("z_outliers.csv", index=False)
    pd.DataFrame(RATIO_WARNINGS).to_csv("ratios_warning.csv", index=False)

    print(f"\n📊 Outlier count: {len(Z_OUTLIERS)}")
    print(f"⚠️ Ratio warnings: {len(RATIO_WARNINGS)}")
    print("📁 Exported: z_outliers.csv + ratios_warning.csv")

    print("\n✅ Summary by Sector & Model:")
    print(df.groupby(["Sector", "Model"])["Status"].value_counts().unstack(fill_value=0))

if __name__ == "__main__":
    main()
